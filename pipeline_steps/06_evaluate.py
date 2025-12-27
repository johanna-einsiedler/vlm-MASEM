#!/usr/bin/env python3
"""Evaluate extraction outputs (factors, correlations, metadata) against ground-truth codings."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl

DATASET_MAP = {
    "tune": "tuning",
    "tuning": "tuning",
    "val": "validation",
    "validation": "validation",
    "eval": "eval",
}


def _resolve_dataset_name(raw_name: str) -> str:
    dataset = DATASET_MAP.get(raw_name)
    if dataset is None:
        raise ValueError(
            f"Unknown dataset '{raw_name}'. Expected one of: {', '.join(DATASET_MAP)}."
        )
    return dataset


def _extract_json_blob(text: str) -> Dict[str, Any]:
    if not isinstance(text, str):
        raise ValueError("Extraction result is not a string.")

    fenced = re.search(r"```json\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fenced:
        return json.loads(fenced.group(1))

    fenced_generic = re.search(r"```\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fenced_generic:
        return json.loads(fenced_generic.group(1))

    inline = re.search(r"(\{.*\})", text, re.DOTALL)
    if inline:
        return json.loads(inline.group(1))

    return json.loads(text)


def _normalize_header(header: Optional[str]) -> Optional[str]:
    if header is None:
        return None
    return str(header).strip().rstrip("*")


def _load_ground_truth_row(excel_path: Path, study: str) -> Dict[str, Any]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = workbook.sheetnames[1]
    sheet = workbook[sheet_name]

    headers = [
        _normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    if "study" not in headers:
        raise ValueError("Missing 'study' column in ground truth sheet.")

    study_index = headers.index("study")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[study_index] == study:
            return {
                header: value
                for header, value in zip(headers, row)
                if header is not None
            }

    raise ValueError(f"Study '{study}' not found in ground truth sheet.")


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _compare_values(extracted: Any, truth: Any, tol: float = 1e-3) -> Dict[str, Any]:
    extracted_value = _to_float(extracted)
    truth_value = _to_float(truth)

    if extracted_value is None and truth_value is None:
        accuracy = 1
    elif extracted_value is None or truth_value is None:
        other = extracted_value if truth_value is None else truth_value
        accuracy = 1 if other == 0 else 0
    else:
        accuracy = 1 if abs(extracted_value - truth_value) <= tol else 0

    return {"extracted": extracted_value, "true": truth_value, "accuracy": accuracy}


def _compare_strings(extracted: Any, truth: Any) -> Dict[str, Any]:
    """Compare string values for metadata fields."""
    extracted_str = str(extracted).strip() if extracted is not None else None
    truth_str = str(truth).strip() if truth is not None else None

    if extracted_str is None and truth_str is None:
        accuracy = 1
    elif extracted_str is None or truth_str is None:
        accuracy = 0
    else:
        # Case-insensitive exact match
        accuracy = 1 if extracted_str.lower() == truth_str.lower() else 0

    return {"extracted": extracted_str, "true": truth_str, "accuracy": accuracy}


def _load_extraction_file(extraction_dir: Path, study_name: str, is_metadata: bool = False) -> Optional[Dict[str, Any]]:
    """Load extraction JSON if it exists, return None otherwise.

    Args:
        extraction_dir: Directory containing extraction files
        study_name: Name of the study (e.g., "parker1993")
        is_metadata: If True, expect {"records": [...]} format instead of samples
    """
    json_path = extraction_dir / f"{study_name}.json"
    if not json_path.exists():
        return None

    try:
        extraction_raw = json.loads(json_path.read_text(encoding="utf-8"))

        # Handle metadata format: {"records": [...]}
        if is_metadata:
            if isinstance(extraction_raw, dict) and "records" in extraction_raw:
                # Convert records to samples format for consistency
                records = extraction_raw.get("records") or []
                # Extract study name from first record if available
                study = records[0].get("study", study_name).rstrip("abc") if records else study_name
                return {
                    "study": study,
                    "samples": records  # Records ARE the samples for metadata
                }
            else:
                return extraction_raw

        # Handle factors/correlations format: {"study": "...", "pages": [{"result": "..."}]}
        if isinstance(extraction_raw, dict):
            # If it has 'study' and 'pages' keys, extract from pages
            if "study" in extraction_raw and "pages" in extraction_raw:
                pages = extraction_raw.get("pages") or []
                if pages:
                    page_result = pages[0].get("result", "")
                    extracted_data = _extract_json_blob(page_result)
                    # Add study name from wrapper
                    if isinstance(extracted_data, dict):
                        extracted_data["study"] = extraction_raw.get("study", study_name)
                    return extracted_data
            # Otherwise, try to extract JSON blob from the content
            elif "result" in extraction_raw:
                return _extract_json_blob(extraction_raw["result"])
            else:
                # Assume it's already the extracted data
                return extraction_raw
        return None
    except Exception as e:
        print(f"Warning: Could not load {json_path}: {e}")
        return None


def run(
    pdf_name: str,
    dataset: str | None = None,
    excel_path: str = "data/ground_truth_codings.xlsx",
    force: bool = False,
) -> list[Path]:
    """Run evaluation on all extraction results (factors, correlations, metadata).

    Args:
        pdf_name: PDF file name (e.g., wise2000.pdf) to locate extraction files
        dataset: Optional dataset name (tuning, validation, eval)
        excel_path: Path to ground truth Excel file
        force: Force re-evaluation even if outputs already exist

    Returns:
        List of paths to evaluation output JSON files
    """
    if dataset:
        dataset = _resolve_dataset_name(dataset)

    study_name = Path(pdf_name).stem

    # Locate all three extraction files
    if dataset:
        factors_dir = Path("data/extraction_factors") / dataset
        correlations_dir = Path("data/extraction_correlations") / dataset
        metadata_dir = Path("data/extraction_metadata") / dataset
    else:
        factors_dir = Path("data/extraction_factors")
        correlations_dir = Path("data/extraction_correlations")
        metadata_dir = Path("data/extraction_metadata")

    # Load extraction data
    factors_data = _load_extraction_file(factors_dir, study_name, is_metadata=False)
    correlations_data = _load_extraction_file(correlations_dir, study_name, is_metadata=False)
    metadata_data = _load_extraction_file(metadata_dir, study_name, is_metadata=True)

    if not factors_data and not correlations_data and not metadata_data:
        print(f"No extraction files found for {study_name}")
        return []

    # Extract samples from each source
    factor_samples = (factors_data.get("samples") or []) if factors_data else []
    correlation_samples = (correlations_data.get("samples") or []) if correlations_data else []
    metadata_samples = (metadata_data.get("samples") or []) if metadata_data else []

    # Determine number of samples
    # Use minimum of non-zero counts to avoid hallucinated duplicates in metadata
    sample_counts = [
        len(factor_samples) if factor_samples else 0,
        len(correlation_samples) if correlation_samples else 0,
        len(metadata_samples) if metadata_samples else 0,
    ]
    non_zero_counts = [c for c in sample_counts if c > 0]
    num_samples = min(non_zero_counts) if non_zero_counts else 0

    if num_samples == 0:
        print(f"No samples found in any extraction for {study_name}")
        return []

    # Get study name from any available source
    study = (
        factors_data.get("study") if factors_data
        else correlations_data.get("study") if correlations_data
        else metadata_data.get("study") if metadata_data
        else study_name
    )

    output_dir = Path("data/evaluation") / dataset if dataset else Path("data/evaluation")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Check if outputs already exist (unless force=True)
    if not force:
        if num_samples > 1:
            expected_files = [
                output_dir / f"{study}{chr(ord('a') + i)}.json"
                for i in range(num_samples)
            ]
        else:
            expected_files = [output_dir / f"{study}.json"]

        # Check if all files exist AND contain all three required sections
        all_complete = True
        for f in expected_files:
            if not f.exists():
                all_complete = False
                break
            try:
                eval_data = json.loads(f.read_text(encoding="utf-8"))
                # Check that all three sections exist
                if not all(key in eval_data for key in ["factor_loadings", "factor_correlations", "metadata"]):
                    all_complete = False
                    break
            except Exception:
                all_complete = False
                break

        if all_complete:
            print(f"Skipping {study}: already evaluated ({len(expected_files)} file(s))")
            return expected_files

    # For multiple samples, try all permutations and keep the best matching
    if num_samples > 1:
        import itertools

        # Generate all possible ground truth names (study + a, b, c, ...)
        gt_names = [f"{study}{chr(ord('a') + i)}" for i in range(num_samples)]

        # Try all permutations of sample-to-ground-truth mappings
        best_permutation = None
        best_total_accuracy = -1

        for perm in itertools.permutations(range(num_samples)):
            total_accuracy = 0
            valid_mapping = True

            # Calculate total accuracy for this permutation
            for sample_idx, gt_idx in enumerate(perm):
                eval_study = gt_names[gt_idx]

                try:
                    ground_truth = _load_ground_truth_row(Path(excel_path), eval_study)
                except ValueError:
                    valid_mapping = False
                    break

                # Get data from each extraction type for this sample
                factor_loadings = (
                    factor_samples[sample_idx].get("factor_loadings", {})
                    if sample_idx < len(factor_samples)
                    else {}
                )
                factor_correlations = (
                    correlation_samples[sample_idx].get("factor_correlations", {})
                    if sample_idx < len(correlation_samples)
                    else {}
                )

                # Calculate accuracy for this sample
                all_comparisons = []
                for key, value in factor_loadings.items():
                    comp = _compare_values(value, ground_truth.get(key))
                    all_comparisons.append(comp["accuracy"])
                for key, value in factor_correlations.items():
                    comp = _compare_values(value, ground_truth.get(key))
                    all_comparisons.append(comp["accuracy"])

                if all_comparisons:
                    sample_accuracy = sum(all_comparisons) / len(all_comparisons)
                    total_accuracy += sample_accuracy

            if valid_mapping and total_accuracy > best_total_accuracy:
                best_total_accuracy = total_accuracy
                best_permutation = perm

        if best_permutation is None:
            print(f"Warning: No valid ground truth mapping found for {study}")
            return []

        # Use the best permutation to create output files
        output_paths: list[Path] = []
        for sample_idx, gt_idx in enumerate(best_permutation):
            eval_study = gt_names[gt_idx]

            # Get data from each extraction type for this sample
            factor_loadings = (
                factor_samples[sample_idx].get("factor_loadings", {})
                if sample_idx < len(factor_samples)
                else {}
            )
            factor_correlations = (
                correlation_samples[sample_idx].get("factor_correlations", {})
                if sample_idx < len(correlation_samples)
                else {}
            )
            # For metadata, match by source_sample_id field
            metadata = {}
            if sample_idx < len(metadata_samples):
                # Get sample_id from factor or correlation data
                sample_id = None
                if sample_idx < len(factor_samples):
                    sample_id = factor_samples[sample_idx].get("sample_id")
                elif sample_idx < len(correlation_samples):
                    sample_id = correlation_samples[sample_idx].get("sample_id")

                # Find matching metadata record by source_sample_id
                if sample_id and metadata_samples:
                    for meta_record in metadata_samples:
                        if meta_record.get("source_sample_id") == sample_id:
                            metadata = meta_record
                            break
                    if not metadata:
                        # Fallback to index-based matching
                        metadata = metadata_samples[sample_idx]
                else:
                    metadata = metadata_samples[sample_idx]

            try:
                ground_truth = _load_ground_truth_row(Path(excel_path), eval_study)
            except ValueError:
                continue

            # Evaluate factor loadings
            evaluated_loadings = {
                key: _compare_values(value, ground_truth.get(key))
                for key, value in factor_loadings.items()
            }

            # Evaluate correlations
            evaluated_correlations = {
                key: _compare_values(value, ground_truth.get(key))
                for key, value in factor_correlations.items()
            }

            # Evaluate metadata (string comparison for most fields)
            evaluated_metadata = {}
            for key, value in metadata.items():
                if key in ["factor_loadings", "factor_correlations"]:
                    # Skip these as they're handled separately
                    continue
                # Check if this is a numeric field in ground truth
                gt_value = ground_truth.get(key)
                if isinstance(gt_value, (int, float)) or (
                    isinstance(gt_value, str) and gt_value.replace(".", "").replace("-", "").isdigit()
                ):
                    # Numeric comparison
                    evaluated_metadata[key] = _compare_values(value, gt_value)
                else:
                    # String comparison
                    evaluated_metadata[key] = _compare_strings(value, gt_value)

            output = {
                "study": eval_study,
                "factor_loadings": evaluated_loadings,
                "factor_correlations": evaluated_correlations,
                "metadata": evaluated_metadata,
            }

            output_path = output_dir / f"{eval_study}.json"
            output_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
            output_paths.append(output_path)

        print(f"Best mapping for {study}: {[f'sample{i}→{gt_names[gt_idx]}' for i, gt_idx in enumerate(best_permutation)]}")
    else:
        # Single sample - use original logic
        output_paths: list[Path] = []
        eval_study = study

        # Get data from each extraction type
        factor_loadings = factor_samples[0].get("factor_loadings", {}) if factor_samples else {}
        factor_correlations = (
            correlation_samples[0].get("factor_correlations", {})
            if correlation_samples
            else {}
        )
        # For single sample, just take the first metadata record
        # (no source_sample_id matching needed)
        metadata = metadata_samples[0] if metadata_samples else {}

        try:
            ground_truth = _load_ground_truth_row(Path(excel_path), eval_study)
        except ValueError:
            print(f"Warning: No ground truth found for {eval_study}")
            return []

        # Evaluate factor loadings
        evaluated_loadings = {
            key: _compare_values(value, ground_truth.get(key))
            for key, value in factor_loadings.items()
        }

        # Evaluate correlations
        evaluated_correlations = {
            key: _compare_values(value, ground_truth.get(key))
            for key, value in factor_correlations.items()
        }

        # Evaluate metadata (string comparison for most fields)
        evaluated_metadata = {}
        for key, value in metadata.items():
            if key in ["factor_loadings", "factor_correlations"]:
                # Skip these as they're handled separately
                continue
            # Check if this is a numeric field in ground truth
            gt_value = ground_truth.get(key)
            if isinstance(gt_value, (int, float)) or (
                isinstance(gt_value, str) and gt_value.replace(".", "").replace("-", "").isdigit()
            ):
                # Numeric comparison
                evaluated_metadata[key] = _compare_values(value, gt_value)
            else:
                # String comparison
                evaluated_metadata[key] = _compare_strings(value, gt_value)

        output = {
            "study": eval_study,
            "factor_loadings": evaluated_loadings,
            "factor_correlations": evaluated_correlations,
            "metadata": evaluated_metadata,
        }

        output_path = output_dir / f"{eval_study}.json"
        output_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
        output_paths.append(output_path)

    return output_paths


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Evaluate extraction results (factors, correlations, metadata) against ground truth."
    )
    parser.add_argument(
        "pdf_name",
        help="PDF file name (e.g., wise2000.pdf) to locate extraction files.",
    )
    parser.add_argument(
        "--dataset",
        help="Optional dataset name: tune, val, eval (or tuning/validation).",
    )
    parser.add_argument(
        "--excel",
        default="data/ground_truth_codings.xlsx",
        help="Path to the ground truth Excel file.",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Force re-evaluation even if output already exists.",
    )
    args = parser.parse_args()

    run(args.pdf_name, dataset=args.dataset, excel_path=args.excel, force=args.force)


if __name__ == "__main__":
    main()
