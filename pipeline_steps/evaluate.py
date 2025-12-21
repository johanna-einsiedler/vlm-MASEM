#!/usr/bin/env python3
"""Evaluate extraction output against ground-truth codings."""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any, Dict, Optional

import openpyxl

DATASET_MAP = {
    "tune": "tuning",
    "tuning": "tuning",
    "val": "validation",
    "validation": "validation",
    "eval": "eval",
}


def _resolve_dataset_name(raw_name: str) -> str:
    dataset = DATASET_MAP.get(raw_name)
    if dataset is None:
        raise ValueError(
            f"Unknown dataset '{raw_name}'. Expected one of: {', '.join(DATASET_MAP)}."
        )
    return dataset


def _extract_json_blob(text: str) -> Dict[str, Any]:
    if not isinstance(text, str):
        raise ValueError("Extraction result is not a string.")

    fenced = re.search(r"```json\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fenced:
        return json.loads(fenced.group(1))

    fenced_generic = re.search(r"```\s*(\{.*?\})\s*```", text, re.DOTALL)
    if fenced_generic:
        return json.loads(fenced_generic.group(1))

    inline = re.search(r"(\{.*\})", text, re.DOTALL)
    if inline:
        return json.loads(inline.group(1))

    return json.loads(text)


def _normalize_header(header: Optional[str]) -> Optional[str]:
    if header is None:
        return None
    return str(header).strip().rstrip("*")


def _load_ground_truth_row(excel_path: Path, study: str) -> Dict[str, Any]:
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    sheet_name = workbook.sheetnames[1]
    sheet = workbook[sheet_name]

    headers = [
        _normalize_header(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))
    ]
    if "study" not in headers:
        raise ValueError("Missing 'study' column in ground truth sheet.")

    study_index = headers.index("study")
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[study_index] == study:
            return {
                header: value
                for header, value in zip(headers, row)
                if header is not None
            }

    raise ValueError(f"Study '{study}' not found in ground truth sheet.")


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return None


def _compare_values(extracted: Any, truth: Any, tol: float = 1e-3) -> Dict[str, Any]:
    extracted_value = _to_float(extracted)
    truth_value = _to_float(truth)

    if extracted_value is None and truth_value is None:
        accuracy = 1
    elif extracted_value is None or truth_value is None:
        other = extracted_value if truth_value is None else truth_value
        accuracy = 1 if other == 0 else 0
    else:
        accuracy = 1 if abs(extracted_value - truth_value) <= tol else 0

    return {"extracted": extracted_value, "true": truth_value, "accuracy": accuracy}


def run(extraction_file: str, dataset: str, excel_path: str) -> list[Path]:
    dataset = _resolve_dataset_name(dataset)
    extraction_path = Path(extraction_file)
    extraction_data = json.loads(extraction_path.read_text(encoding="utf-8"))

    study = extraction_data.get("study")
    pages = extraction_data.get("pages") or []
    if not study or not pages:
        raise ValueError("Extraction JSON missing 'study' or 'pages'.")

    page_info = pages[0]
    page_number = page_info.get("number")
    extracted_payload = _extract_json_blob(page_info.get("result", ""))

    samples = extracted_payload.get("samples") or []
    if not samples:
        raise ValueError("No samples found in extraction payload.")

    output_dir = Path("data/evaluation") / dataset
    output_dir.mkdir(parents=True, exist_ok=True)

    output_paths: list[Path] = []
    for index, sample in enumerate(samples):
        suffix = chr(ord("a") + index)
        eval_study = f"{study}{suffix}"
        factor_loadings = sample.get("factor_loadings") or {}
        factor_correlations = sample.get("factor_correlations") or {}

        try:
            ground_truth = _load_ground_truth_row(Path(excel_path), eval_study)
        except ValueError:
            continue

        evaluated_loadings = {
            key: _compare_values(value, ground_truth.get(key))
            for key, value in factor_loadings.items()
        }
        evaluated_correlations = {
            key: _compare_values(value, ground_truth.get(key))
            for key, value in factor_correlations.items()
        }

        output = {
            "study": eval_study,
            "page": page_number,
            "factor_loadings": evaluated_loadings,
            "factor_correlations": evaluated_correlations,
        }

        output_path = output_dir / f"{eval_study}.json"
        output_path.write_text(json.dumps(output, indent=2), encoding="utf-8")
        output_paths.append(output_path)

    print(json.dumps({"study": study, "pages": pages}, indent=2))
    return output_paths


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Evaluate extraction results against ground truth."
    )
    parser.add_argument(
        "extraction_file",
        help="Path to the extraction JSON file (e.g., data/extraction/erni1997.json).",
    )
    parser.add_argument(
        "dataset",
        help="Dataset name: tune, val, eval (or tuning/validation).",
    )
    parser.add_argument(
        "--excel",
        default="data/ground_truth_codings.xlsx",
        help="Path to the ground truth Excel file.",
    )
    args = parser.parse_args()

    run(args.extraction_file, args.dataset, args.excel)


if __name__ == "__main__":
    main()
